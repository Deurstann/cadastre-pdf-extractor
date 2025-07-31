# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
Interface Streamlit Pro - Extracteur Cadastral Français
Version client - Interface moderne et simplifiée
"""

import streamlit as st
import tempfile
import pandas as pd
from pathlib import Path
from cadastrePdfExtractor import CadastralPdfExtractor
import io
import logging

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration de la page
st.set_page_config(
    page_title="Extracteur Cadastral Pro",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# CSS moderne
st.markdown(
    """
<style>
    .main-header {
        font-size: 3rem;
        font-weight: 700;
        background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.3rem;
        color: #64748b;
        text-align: center;
        margin-bottom: 2.5rem;
        font-weight: 400;
    }
    .upload-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2.5rem;
        border-radius: 15px;
        margin: 2rem 0;
        text-align: center;
    }
    .upload-text {
        color: white;
        font-size: 1.2rem;
        font-weight: 500;
        margin-bottom: 1.5rem;
    }
    .results-container {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        padding: 2rem;
        border-radius: 15px;
        border: 2px solid #0ea5e9;
        margin: 2rem 0;
    }
    .download-container {
        background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
        padding: 2rem;
        border-radius: 15px;
        border: 2px solid #22c55e;
        margin: 2rem 0;
    }
    .metric-box {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        text-align: center;
        border: 1px solid #e2e8f0;
    }
    .metric-number {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1e40af;
        margin-bottom: 0.5rem;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #64748b;
        font-weight: 500;
    }
    .stButton > button {
        width: 100%;
        border-radius: 10px;
        font-weight: 600;
        font-size: 1.1rem;
        padding: 0.75rem;
        border: none;
        transition: all 0.3s ease;
    }
    .feature-badge {
        background: #f1f5f9;
        color: #475569;
        padding: 0.5rem 1rem;
        border-radius: 25px;
        font-size: 0.9rem;
        font-weight: 500;
        margin: 0.25rem;
        display: inline-block;
    }
    .footer {
        background: #f8f9fa;
        padding: 2rem;
        text-align: center;
        margin-top: 3rem;
        border-top: 1px solid #e9ecef;
        color: #6c757d;
        font-size: 0.9rem;
    }
    .footer-logo {
        font-weight: 600;
        color: #495057;
        font-size: 1.1rem;
    }
</style>
""",
    unsafe_allow_html=True,
)


def create_excel_download(df):
    """Crée un fichier Excel téléchargeable."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Données Cadastrales")

        # Formatter le fichier Excel
        workbook = writer.book
        worksheet = writer.sheets["Données Cadastrales"]

        # Style pour les en-têtes (pour compatibilité)
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Appliquer le style aux en-têtes
            for col_num, _ in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                worksheet.column_dimensions[cell.column_letter].width = 15
        except ImportError:
            pass

    return output.getvalue()


def main():

    cadatral_pdf_extractor = CadastralPdfExtractor()

    # En-tête moderne
    st.markdown(
        '<h1 class="main-header">Extracteur Cadastral Pro</h1>', unsafe_allow_html=True
    )
    st.markdown(
        '<p class="sub-header">Solution professionnelle d\'extraction de données cadastrales françaises</p>',
        unsafe_allow_html=True,
    )

    if "extraction_results" not in st.session_state:
        st.session_state.extraction_results = None
    if "processed_files" not in st.session_state:
        st.session_state.processed_files = []
    if "current_file_hash" not in st.session_state:
        st.session_state.current_file_hash = None

    # Badges des fonctionnalités
    st.markdown(
        """
    <div style="text-align: center; margin-bottom: 2rem;">
        <span class="feature-badge">Export Excel & CSV</span>
        <span class="feature-badge">Traitement Rapide</span>
        <span class="feature-badge">Haute Précision</span>
    </div>
    """,
        unsafe_allow_html=True,
    )

    # Section upload
    st.markdown(
        """
    <div class="upload-container">
        <div class="upload-text">
            Glissez-déposez vos fichiers PDF cadastraux ou cliquez pour les sélectionner
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    uploaded_files = st.file_uploader(
        label="none",
        type=["pdf"],
        accept_multiple_files=True,
        help="Formats supportés: PDF | Taille max: 200MB par fichier",
        label_visibility="collapsed",
    )

    if uploaded_files:
        current_files_hash = hash(
            tuple(f.name + str(len(f.getvalue())) for f in uploaded_files)
        )

        if st.session_state.current_file_hash != current_files_hash:
            st.session_state.extraction_results = None
            st.session_state.processed_files = []
            st.session_state.current_file_hash = current_files_hash
            st.rerun()
    elif st.session_state.extraction_results is not None:
        st.session_state.extraction_results = None
        st.session_state.processed_files = []
        st.session_state.current_file_hash = None
        st.rerun()

    with st.expander("Options avancées", expanded=False):
        debug_mode = st.checkbox(
            "Mode diagnostic (pour le support technique)", value=False
        )
        col1, col2 = st.columns(2)
        with col1:
            export_format = st.selectbox(
                "Format d'export principal", ["Excel (.xlsx)", "CSV (.csv)"], index=0
            )
        with col2:
            include_empty = st.checkbox("Inclure les champs vides", value=True)

    if uploaded_files:
        st.success(f"{len(uploaded_files)} fichier(s) chargé(s) avec succès")

        if st.session_state.extraction_results:
            st.info(
                "💡 Résultats d'extraction disponibles ci-dessous. Cliquez sur 'Démarrer l'extraction' pour retraiter les fichiers."
            )

        # Informations sur les fichiers
        with st.expander("Détails des fichiers", expanded=False):
            for file in uploaded_files:
                file_size = len(file.getvalue()) / 1024 / 1024  # MB
                st.write(f"• **{file.name}** - {file_size:.1f} MB")

        button_text = (
            "Retraiter les fichiers"
            if st.session_state.extraction_results
            else "Démarrer l'extraction"
        )

        if st.button(button_text, type="primary", use_container_width=True):

            st.session_state.extraction_results = None
            st.session_state.processed_files = []

            with st.status("Traitement en cours...", expanded=True) as status:

                # Créer un répertoire temporaire
                with tempfile.TemporaryDirectory() as temp_dir:
                    temp_path = Path(temp_dir)
                    input_dir = temp_path / "input"
                    output_dir = temp_path / "output"
                    input_dir.mkdir()
                    output_dir.mkdir()

                    st.write("Préparation des fichiers...")

                    # Sauvegarder les fichiers
                    saved_files: list = []
                    for uploaded_file in uploaded_files:
                        file_path = input_dir / uploaded_file.name
                        with open(file_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())
                        saved_files.append(file_path)

                    try:

                        all_properties = []

                        progress_bar = st.progress(0)

                        for i, pdf_file in enumerate(saved_files):
                            st.write(f"Traitement: {pdf_file.name}")
                            progress_bar.progress((i + 1) / len(saved_files))

                            properties = cadatral_pdf_extractor.extract_cadatral_info_from_pdf(pdf_file)
                            all_properties.extend(properties)

                        if all_properties:
                            st.session_state.extraction_results = all_properties
                            st.session_state.processed_files = [
                                f.name for f in uploaded_files
                            ]
                            st.success(
                                f"Extraction terminée ! {len(all_properties)} propriétés extraites."
                            )
                            st.rerun()
                        else:
                            st.warning(
                                "Aucune propriété extraite. Vérifiez vos fichiers PDF."
                            )
                            st.session_state.extraction_results = []

                        status.update(label="Extraction terminée!", state="complete")

                    except Exception as e:
                        st.error(f"Erreur lors du traitement: {e}")
                        st.session_state.extraction_results = []
                        status.update(
                            label="Erreur pendant l'extraction", state="error"
                        )

    if st.session_state.extraction_results:
        # Vérification de cohérence des résultats
        if uploaded_files:
            current_files = [f.name for f in uploaded_files]
            if st.session_state.processed_files != current_files:
                st.warning(
                    "Attention : Les résultats affichés proviennent de fichiers différents."
                )
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("🧹 Nettoyer et retraiter", type="secondary"):
                        st.session_state.extraction_results = None
                        st.session_state.processed_files = []
                        st.session_state.current_file_hash = None
                        st.rerun()
                with col2:
                    if st.button("📋 Garder les résultats", type="primary"):
                        st.info(
                            "Résultats conservés. Vous pouvez les télécharger ci-dessous."
                        )
        # Préparation des données
        df_display = pd.DataFrame(st.session_state.extraction_results)

        # Conteneur des résultats
        st.markdown(
            """
        <div class="results-container">
            <h3 style="color: #0ea5e9; margin-bottom: 1.5rem;">Résultats de l'extraction</h3>
        </div>
        """,
            unsafe_allow_html=True,
        )

        # Métriques
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            unique_properties = len(df_display[df_display["id"].notna()]["id"].unique())
            st.markdown(
                f"""
            <div class="metric-box">
                <div class="metric-number">{unique_properties}</div>
                <div class="metric-label">Propriétés</div>
            </div>
            """,
                unsafe_allow_html=True,
            )

        with col2:
            unique_owners = len(df_display[df_display["numero_majic"].notna()]["numero_majic"].unique())
            st.markdown(
                f"""
            <div class="metric-box">
                <div class="metric-number">{unique_owners}</div>
                <div class="metric-label">Propriétaires</div>
            </div>
            """,
                unsafe_allow_html=True,
            )

        with col3:
            files_processed = len(df_display["fichier_source"].unique())
            st.markdown(
                f"""
            <div class="metric-box">
                <div class="metric-number">{files_processed}</div>
                <div class="metric-label">Fichiers</div>
            </div>
            """,
                unsafe_allow_html=True,
            )

        with col4:
            completion_rate = df_display["nom"].notna().sum() / len(df_display) * 100
            st.markdown(
                f"""
            <div class="metric-box">
                <div class="metric-number">{completion_rate:.0f}%</div>
                <div class="metric-label">Complétude</div>
            </div>
            """,
                unsafe_allow_html=True,
            )

        st.markdown("<br>", unsafe_allow_html=True)

        # Aperçu des données
        st.markdown("### Aperçu des données")
        st.dataframe(df_display, use_container_width=True, height=400)

        # Section téléchargement
        st.markdown(
            """
        <div class="download-container">
            <h3 style="color: #22c55e; margin-bottom: 1.5rem;">Télécharger vos données</h3>
        </div>
        """,
            unsafe_allow_html=True,
        )

        # Préparer les données pour export
        df_export = df_display.copy()

        # Renommer les colonnes pour l'export
        column_mapping = {
            "department": "Département",
            "commune": "Commune",
            "prefixe": "Préfixe",
            "section": "Section",
            "plan_number": "Numéro",
            "contenance_ha": "Contenance HA",
            "contenance_a": "Contenance A",
            "contenance_ca": "Contenance CA",
            "droit_reel": "Droit réel",
            "designation_parcelle": "Designation Parcelle",
            "nom": "Nom Propri",
            "prenom": "Prénom Propri",
            "numero_majic": "N°MAJIC",
            "voie": "Voie",
            "code_postal": "CP",
            "ville": "Ville",
            "id": "ID",
            "fichier_source": "Fichier source",
        }

        df_export = df_export.rename(columns=column_mapping)

        # Boutons de téléchargement
        col1, col2 = st.columns(2)

        with col1:
            # Export Excel
            excel_data = create_excel_download(df_export)
            st.download_button(
                label="Télécharger Excel",
                data=excel_data,
                file_name="extraction_cadastrale.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col2:
            # Export CSV
            csv_data = df_export.to_csv(index=False, sep=";", encoding="utf-8-sig")
            st.download_button(
                label="Télécharger CSV",
                data=csv_data,
                file_name="extraction_cadastrale.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # Informations techniques (séparées des autres contextes)
    if st.session_state.extraction_results:
        with st.expander("Informations techniques", expanded=False):
            st.markdown(
                """
            **Format des données:**
            - **Excel**: Format .xlsx avec mise en forme automatique
            - **CSV**: Séparateur point-virgule (;), encodage UTF-8

            **Colonnes extraites:**
            - Informations parcellaires: Département, Commune, Préfixe, Section, Numéro
            - Contenance détaillée: HA (hectares), A (ares), CA (centiares)
            - Propriétaire: Nom, Prénom, N°MAJIC, Adresse complète
            - Identification unique: ID 14 caractères
            """
            )
    elif not uploaded_files:
        if st.session_state.extraction_results:
            st.info(
                "Résultats précédents encore affichés. Rechargez des fichiers pour retraiter."
            )
            if st.button("🧹 Nettoyer tous les résultats", type="secondary"):
                st.session_state.extraction_results = None
                st.session_state.processed_files = []
                st.session_state.current_file_hash = None
                st.rerun()
        else:
            st.info(
                "Sélectionnez un ou plusieurs fichiers PDF pour commencer l'extraction."
            )

        # Informations d'aide
        with st.expander("Comment utiliser cet outil", expanded=False):
            st.markdown(
                """
            ### Guide d'utilisation

            1. **Sélectionnez vos fichiers PDF** cadastraux français
            2. **Cliquez sur "Démarrer l'extraction"** pour lancer le traitement
            3. **Consultez les résultats** dans le tableau interactif
            4. **Téléchargez vos données** en Excel ou CSV

            ### Types de données extraites
            - **Informations parcellaires** (département, commune, section, numéro)
            - **Contenance détaillée** (hectares, ares, centiares)
            - **Propriétaires** (nom, prénom, adresse)
            - **Identifiants** (N°MAJIC, ID unique)

            ### Formats supportés
            - **PDF natifs** (générés numériquement)
            """
            )

    # Footer avec copyright Level Up
    st.markdown(
        """
    <div class="footer">
        <div class="footer-logo">Level Up</div>
        <div>© 2024 Level Up. Tous droits réservés.</div>
        <div style="margin-top: 0.5rem; font-size: 0.8rem;">
            Solution d'extraction cadastrale professionnelle
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
